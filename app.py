from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import sqlite3
import random
import os
import pandas as pd

app = Flask(__name__)
app.secret_key = "sirs_official_system_2026"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'students.db')

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY, 
                    username TEXT, 
                    password TEXT, 
                    role TEXT)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS records (
                    student_id TEXT PRIMARY KEY, 
                    name TEXT, 
                    section TEXT, 
                    gender TEXT,
                    course TEXT,
                    ww1 REAL DEFAULT 0, ww2 REAL DEFAULT 0, ww3 REAL DEFAULT 0,
                    pt1 REAL DEFAULT 0, pt2 REAL DEFAULT 0, pt3 REAL DEFAULT 0,
                    pre REAL DEFAULT 0, mid REAL DEFAULT 0, fin REAL DEFAULT 0,
                    gpa REAL DEFAULT 0)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS grades (
                    student_id TEXT,
                    subject TEXT,
                    semester TEXT,
                    written REAL DEFAULT 0,
                    pt REAL DEFAULT 0,
                    pre REAL DEFAULT 0,
                    mid REAL DEFAULT 0,
                    fin REAL DEFAULT 0,
                    final_point REAL DEFAULT 0,
                    PRIMARY KEY (student_id, subject, semester))''')
    
    c.execute("INSERT OR IGNORE INTO users (username, password, role) VALUES (?, ?, ?)", 
              ('admin', 'password123', 'teacher'))
    conn.commit()
    conn.close()

def generate_unique_id():
    return str(random.randint(100000, 999999))

@app.route('/')
def index():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    username = request.form.get('username')
    password = request.form.get('password')
    
    if username == 'admin' and password == 'password123':
        session['role'] = 'teacher'
        session['user'] = 'Admin'
        return redirect(url_for('teacher_dashboard'))
    
    session['role'] = 'student'
    session['user'] = username
    return redirect(url_for('index'))

@app.route('/teacher')
def teacher_dashboard():
    if session.get('role') != 'teacher':
        return redirect(url_for('index'))
        
    view = request.args.get('view', 'home')
    sec_filter = request.args.get('section', '')
    sub_filter = request.args.get('subject', 'IPT') 
    sem_filter = request.args.get('semester', '1st')
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # Dashboard Home Stats
    c.execute("SELECT COUNT(*) FROM records")
    total_students = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM records WHERE gender = 'Boy'")
    boys_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM records WHERE gender = 'Girl'")
    girls_count = c.fetchone()[0]
    
    c.execute("SELECT section, COUNT(*) FROM records GROUP BY section")
    section_folders = c.fetchall()
    c.execute("SELECT DISTINCT section FROM records")
    sections = [row[0] for row in c.fetchall()]
    
    students = []
    stats = {'total': 0, 'on_track': 0, 'at_risk': 0}
    
    if view == 'grades':
        # Makuha ang grades base sa subject at semester
        c.execute('''
            SELECT r.student_id, r.name, g.written, g.pt, g.pre, g.mid, g.fin, g.final_point, r.section, r.gender, r.course
            FROM records r
            LEFT JOIN grades g ON r.student_id = g.student_id 
            AND g.subject = ? AND g.semester = ?
        ''', (sub_filter, sem_filter))
        raw_students = c.fetchall()
        
        # dynamic GPA processing across all subjects for each student record row
        processed_students = []
        for s in raw_students:
            s_list = list(s)
            sid = s_list[0]
            
            # Fetch all graded final points to evaluate true composite cumulative GPA
            c_gpa = sqlite3.connect(DB_PATH).cursor()
            c_gpa.execute("SELECT final_point FROM grades WHERE student_id = ? AND final_point > 0", (sid,))
            all_grades = [row[0] for row in c_gpa.fetchall()]
            
            calc_gpa = round(sum(all_grades) / len(all_grades), 2) if all_grades else 0.00
            s_list.append(calc_gpa) # Index [11]: Dynamic GPA tracking token
            processed_students.append(s_list)
            
        students = processed_students
        
        # Calculate stats para sa grades view
        stats['total'] = len(raw_students)
        stats['on_track'] = sum(1 for s in raw_students if s[7] is not None and s[7] > 0 and s[7] <= 2.75)
        stats['at_risk'] = sum(1 for s in raw_students if s[7] is not None and (s[7] >= 3.00 or s[7] == 5.00))

    elif sec_filter:
        # Kukunin ang records pati ang connected grade/final_point para sa View by Section card view
        c.execute('''
            SELECT r.student_id, r.name, r.section, r.gender, r.course, 
                   r.ww1, r.ww2, r.ww3, r.pt1, r.pt2, r.pt3, r.pre, r.mid, r.fin,
                   g.final_point
            FROM records r
            LEFT JOIN grades g ON r.student_id = g.student_id AND g.subject = ? AND g.semester = ?
            WHERE r.section = ?
        ''', (sub_filter, sem_filter, sec_filter))
        raw_students = c.fetchall()
        
        processed_students = []
        for s in raw_students:
            s_list = list(s)
            sid = s_list[0]
            
            c_gpa = sqlite3.connect(DB_PATH).cursor()
            c_gpa.execute("SELECT final_point FROM grades WHERE student_id = ? AND final_point > 0", (sid,))
            all_grades = [row[0] for row in c_gpa.fetchall()]
            
            calc_gpa = round(sum(all_grades) / len(all_grades), 2) if all_grades else 0.00
            s_list.append(calc_gpa) # Index [15]: Appended dynamic overall GPA
            processed_students.append(s_list)
            
        students = processed_students
        
        # Calculate stats base sa 1.00 - 5.00 rules mo para sa section folders
        stats['total'] = len(raw_students)
        stats['on_track'] = sum(1 for s in raw_students if s[14] is not None and s[14] > 0 and s[14] <= 2.75)
        stats['at_risk'] = sum(1 for s in raw_students if s[14] is not None and (s[14] >= 3.00 or s[14] == 5.00))
    else:
        # Pangkalahatang view ng records
        c.execute('''
            SELECT r.student_id, r.name, r.section, r.gender, r.course, 
                   r.ww1, r.ww2, r.ww3, r.pt1, r.pt2, r.pt3, r.pre, r.mid, r.fin,
                   g.final_point
            FROM records r
            LEFT JOIN grades g ON r.student_id = g.student_id AND g.subject = ? AND g.semester = ?
        ''', (sub_filter, sem_filter))
        raw_students = c.fetchall()
        
        processed_students = []
        for s in raw_students:
            s_list = list(s)
            sid = s_list[0]
            
            c_gpa = sqlite3.connect(DB_PATH).cursor()
            c_gpa.execute("SELECT final_point FROM grades WHERE student_id = ? AND final_point > 0", (sid,))
            all_grades = [row[0] for row in c_gpa.fetchall()]
            
            calc_gpa = round(sum(all_grades) / len(all_grades), 2) if all_grades else 0.00
            s_list.append(calc_gpa) # Index [15]: Appended dynamic overall GPA
            processed_students.append(s_list)
            
        students = processed_students
        
        stats['total'] = len(raw_students)
        stats['on_track'] = sum(1 for s in raw_students if s[14] is not None and s[14] > 0 and s[14] <= 2.75)
        stats['at_risk'] = sum(1 for s in raw_students if s[14] is not None and (s[14] >= 3.00 or s[14] == 5.00))

    new_id = generate_unique_id()
    conn.close()

    return render_template('teacher_dashboard.html',
        view=view, total=total_students, boys=boys_count, girls=girls_count,
        sections=sections, section_folders=section_folders, students=students,
        current_sec=sec_filter, current_sub=sub_filter, current_sem=sem_filter,
        stats=stats, new_student_id=new_id
    )

@app.route('/save_grades', methods=['POST'])
def save_grades():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    subject = request.form.get('current_subject')
    semester = request.form.get('current_semester')
    
    # HPS (Over) Scores
    m_w = float(request.form.get('max_written') or 100)
    m_p = float(request.form.get('max_pt') or 100)
    m_pre = float(request.form.get('max_pre') or 100)
    m_mid = float(request.form.get('max_mid') or 100)
    m_fin = float(request.form.get('max_fin') or 100)

    student_ids = request.form.getlist('student_id[]')
    written_list = request.form.getlist('written[]')
    pt_list = request.form.getlist('pt[]')
    pre_list = request.form.getlist('pre[]')
    mid_list = request.form.getlist('mid[]')
    fin_list = request.form.getlist('fin[]')

    for i in range(len(student_ids)):
        sid = student_ids[i]
        w = float(written_list[i] or 0)
        p = float(pt_list[i] or 0)
        pr = float(pre_list[i] or 0)
        mi = float(mid_list[i] or 0)
        fi = float(fin_list[i] or 0)

        # Computation (20% Written, 40% PT, 40% Exams [10-10-20])
        w_pct = (w / m_w * 100) if m_w > 0 else 0
        p_pct = (p / m_p * 100) if m_p > 0 else 0
        pre_pct = (pr / m_pre * 100) if m_pre > 0 else 0
        mid_pct = (mi / m_mid * 100) if m_mid > 0 else 0
        fin_pct = (fi / m_fin * 100) if m_fin > 0 else 0

        raw = (w_pct * 0.20) + (p_pct * 0.40) + (pre_pct * 0.10) + (mid_pct * 0.10) + (fin_pct * 0.20)
        
        # College Transmutation: 75=3.0, 100=1.0
        if raw >= 75:
            final_point = round(max(1.0, 1.0 + (100 - raw) * (2.0 / 25.0)), 2)
        else:
            final_point = 5.00 if raw > 0 else 0

        c.execute('''INSERT INTO grades (student_id, subject, semester, written, pt, pre, mid, fin, final_point)
                     VALUES (?,?,?,?,?,?,?,?,?)
                     ON CONFLICT(student_id, subject, semester) DO UPDATE SET
                     written=excluded.written, pt=excluded.pt, pre=excluded.pre, 
                     mid=excluded.mid, fin=excluded.fin, final_point=excluded.final_point''', 
                  (sid, subject, semester, w, p, pr, mi, fi, final_point))
    
    conn.commit()
    conn.close()
    flash("Grades computed and saved!")
    return redirect(url_for('teacher_dashboard', view='grades', subject=subject, semester=semester))

@app.route('/export_grades/<section>')
def export_grades(section):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT r.student_id, r.name, g.subject, g.written, g.pt, g.pre, g.mid, g.fin, g.final_point
        FROM records r
        LEFT JOIN grades g ON r.student_id = g.student_id
        WHERE r.section = ?
    """, (section,))
    records = cursor.fetchall()
    conn.close()

    if not records:
        return f"<h1>No student records found for section {section}</h1>"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{section} Grades"
    ws.views.sheetView[0].showGridLines = True
    
    maroon_fill = PatternFill(start_color="4A0E17", end_color="4A0E17", fill_type="solid")
    pass_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    white_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    headers = ["Student ID", "Full Name", "Subject Title", "Written (20%)", "Performance (40%)", "Prelim (10%)", "Midterm (10%)", "Finals (20%)", "Final Rating", "Remarks"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = maroon_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 2
    for row in records:
        ws.cell(row=current_row, column=1, value=row['student_id'])
        ws.cell(row=current_row, column=2, value=row['name'])
        ws.cell(row=current_row, column=3, value=row['subject'] or "N/A")
        ws.cell(row=current_row, column=4, value=row['written'])
        ws.cell(row=current_row, column=5, value=row['pt'])
        ws.cell(row=current_row, column=6, value=row['pre'])
        ws.cell(row=current_row, column=7, value=row['mid'])
        ws.cell(row=current_row, column=8, value=row['fin'])
        
        final_rating = row['final_point']
        rcell = ws.cell(row=current_row, column=9, value=final_rating if final_rating else 0.00)
        rcell.number_format = '0.00'
        
        rem_cell = ws.cell(row=current_row, column=10)
        if final_rating and final_rating <= 3.00:
            rem_cell.value = "PASSED"
            rem_cell.fill = pass_fill
        else:
            rem_cell.value = "FAILED"
            
        current_row += 1

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 16

    filename = f"SIRS_Grades_{section}.xlsx"
    file_path = os.path.join(BASE_DIR, 'static', filename)
    wb.save(file_path)
    
    return send_file(file_path, as_attachment=True)

@app.route('/export_excel')
def export_excel():
    conn = sqlite3.connect(DB_PATH)
    query = "SELECT r.*, g.subject, g.semester, g.final_point FROM records r LEFT JOIN grades g ON r.student_id = g.student_id"
    df = pd.read_sql_query(query, conn)
    conn.close()
    file_path = os.path.join(BASE_DIR, "SIRS_Full_Report.xlsx")
    df.to_excel(file_path, index=False)
    return send_file(file_path, as_attachment=True)

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    file = request.files.get('file')
    if file and file.filename.endswith(('.xlsx', '.xls')):
        try:
            df = pd.read_excel(file)
            flash(f"Successfully imported {len(df)} records!")
        except Exception as e:
            flash(f"Error: {str(e)}")
    return redirect(url_for('teacher_dashboard', view='grades'))

@app.route('/add_student', methods=['POST'])
def add_student():
    student_id, name, section, gender = request.form.get('student_id'), request.form.get('name'), request.form.get('section'), request.form.get('gender')
    course = request.form.get('course', 'Academic') 
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("INSERT INTO records (student_id, name, section, gender, course) VALUES (?,?,?,?,?)", 
                     (student_id, name, section, gender, course))
        conn.commit()
        flash(f"Student {name} added!")
    except: 
        flash("Error: ID already exists.")
    finally: 
        conn.close()
    return redirect(url_for('teacher_dashboard', view='home'))

@app.route('/delete_student/<sid>')
def delete_student(sid):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM records WHERE student_id = ?", (sid,))
    conn.commit()
    conn.close()
    return redirect(url_for('teacher_dashboard', view='sections'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/student_profile/<sid>')
def student_profile(sid):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("SELECT student_id, name, section, course FROM records WHERE student_id = ?", (sid,))
    student = cursor.fetchone()
    
    cursor.execute("SELECT * FROM grades WHERE student_id = ?", (sid,))
    grades = cursor.fetchall()
    
    cursor.execute("SELECT AVG(final_point) FROM grades WHERE student_id = ? AND final_point IS NOT NULL AND final_point > 0", (sid,))
    row = cursor.fetchone()
    overall_gpa = row[0] if row[0] is not None else 0.00
    
    conn.close()
    
    if not student:
        flash("Student not found!", "danger")
        return redirect(url_for('teacher_dashboard'))
        
    return render_template('student_profile.html', student=student, grades=grades, overall_gpa=overall_gpa)

if __name__ == '__main__':
    init_db()
    app.run(debug=True)